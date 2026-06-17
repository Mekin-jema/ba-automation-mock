#!/usr/bin/env python3
"""
Report Generator - Create pivot report from CSV outputs
Converts regional query CSVs into Excel pivot format (Regions as rows, Dates as columns)
Matches the CBU Daily Stand-Up report format
"""

from pathlib import Path

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False


def load_csv(csv_path: Path) -> pd.DataFrame:
    """Load CSV file into DataFrame"""
    return pd.read_csv(csv_path)


def is_regional_query(df: pd.DataFrame) -> bool:
    """Check if DataFrame contains regional data"""
    regional_keywords = ["region", "exec_region", "ds_exec_region"]
    columns_lower = [col.lower() for col in df.columns]
    return any(kw in " ".join(columns_lower) for kw in regional_keywords)


def create_regional_pivot(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create pivot table: Regions as rows, Dates as columns
    Format: like the CBU Daily Stand-Up report
    """
    
    # Find column names
    region_col = None
    date_col = None
    metric_col = None
    
    for col in df.columns:
        col_lower = col.lower()
        if "region" in col_lower or "exec_region" in col_lower:
            region_col = col
        elif "date" in col_lower or "id_date" in col_lower:
            date_col = col
    
    # If no date col, use first non-region column as metric
    if not metric_col:
        for col in df.columns:
            if col != region_col:
                metric_col = col
                break
    
    if not region_col or not date_col:
        return df  # Can't pivot
    
    try:
        # Pivot: regions as rows, dates as columns
        pivot_df = df.pivot_table(
            index=region_col,
            columns=date_col,
            values=metric_col,
            aggfunc='first'
        )
        return pivot_df
    except Exception as e:
        print(f"  Pivot warning: {e}")
        return df


def generate_regional_report(output_dir: Path = Path("outputs"), 
                            report_path: Path = Path("DAILY_REGIONAL_REPORT.xlsx")) -> Path:
    """Generate Excel report with regional pivots"""
    
    if not HAS_EXCEL:
        print("❌ pandas and openpyxl required. Install: pip install pandas openpyxl")
        return None
    
    print(f"\n📊 Generating Regional Pivot Report from {output_dir}...\n")
    
    # Find all CSV files
    csv_files = sorted(output_dir.glob("query_*.csv"))
    
    if not csv_files:
        print("❌ No query CSV files found in outputs/")
        return None
    
    # Create Excel writer
    with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
        
        regional_count = 0
        
        for csv_file in csv_files:
            try:
                df = load_csv(csv_file)
                query_num = csv_file.stem.replace("query_", "")
                
                # Check if it has regional data
                if is_regional_query(df):
                    print(f"✓ Query {query_num}: Pivoting regional data...")
                    
                    # Create pivot
                    pivoted = create_regional_pivot(df)
                    
                    # Write to Excel
                    sheet_name = f"Q{query_num}_Regional"
                    pivoted.to_excel(writer, sheet_name=sheet_name)
                    
                    print(f"  → Sheet: {sheet_name} ({len(pivoted)} regions)")
                    regional_count += 1
                else:
                    # Non-regional query - write as-is
                    sheet_name = f"Q{query_num}"
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"✓ Query {query_num}: {sheet_name}")
                    
            except Exception as e:
                print(f"✗ Query {csv_file.stem}: {e}")
        
        print(f"\n✅ Created {regional_count} regional pivot sheets")
    
    # Format Excel
    try:
        wb = openpyxl.load_workbook(report_path)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            
            # Format headers
            for cell in ws[1]:
                if cell.value:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
            
            # Format data cells and auto-width
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for idx, cell in enumerate(column):
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                        if idx > 0:  # Skip header
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="center")
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        wb.save(report_path)
        print(f"✓ Formatted and saved: {report_path}\n")
    except Exception as e:
        print(f"⚠️  Format warning: {e}\n")
    
    return report_path


if __name__ == "__main__":
    # Generate regional pivot report
    report = generate_regional_report(
        output_dir=Path("outputs"),
        report_path=Path("DAILY_REGIONAL_REPORT.xlsx")
    )
    
    if report and report.exists():
        print(f"📂 Report ready: {report}")
        print("   Open in Excel to view regional pivots!")
