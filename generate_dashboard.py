#!/usr/bin/env python3
"""
Generate Dashboard - Merge query results into the CBU Stand-Up Report format.
Pivots the SQL output data so that dates are columns, matching the layout of real-sample-report.csv.
Accumulates historical columns day-by-day.
"""

from pathlib import Path
from datetime import datetime
import pandas as pd
import numpy as np

# Define paths
WORKSPACE_DIR = Path(__file__).parent
OUTPUTS_DIR = WORKSPACE_DIR / "outputs"
TEMPLATE_PATH = WORKSPACE_DIR / "real-sample-report.csv"
OUTPUT_REPORT_PATH = OUTPUTS_DIR / "real-sample-report-pivoted.csv"

# Global data cache
df_cache = {}

def get_query_df(query_num: int) -> pd.DataFrame:
    path = OUTPUTS_DIR / f"query_{query_num:02d}.csv"
    if not path.exists():
        return None
    if query_num not in df_cache:
        df_cache[query_num] = pd.read_csv(path)
    return df_cache[query_num]

def get_date_col(df: pd.DataFrame) -> str:
    for col in df.columns:
        if col.lower() in ['id_date', 'event_date']:
            return col
    return df.columns[0]

def clean_region_name(val) -> str:
    if not isinstance(val, str):
        return ""
    val = val.strip()
    if "." in val:
        val = val.split(".", 1)[1].strip()
    if val.lower() == "uknown" or val.lower() == "unknown":
        val = "Unknown"
    return val

def get_unique_dates() -> list[int]:
    dates = set()
    for query_num in range(1, 20):
        df = get_query_df(query_num)
        if df is not None and len(df) > 0:
            date_col = get_date_col(df)
            for d in df[date_col].dropna().unique():
                try:
                    d_str = str(int(float(d)))
                    if len(d_str) == 8:
                        dates.add(int(d_str))
                except (ValueError, TypeError):
                    pass
    return sorted(list(dates))

def format_date_headers(date_int: int) -> tuple[str, str]:
    try:
        dt = datetime.strptime(str(date_int), "%Y%m%d")
        day_of_week = dt.strftime("%a")
        date_label = f"{dt.day}-{dt.strftime('%b')}"
        return day_of_week, date_label
    except Exception:
        return "", str(date_int)

def extract_metric(query_num: int, date_int: int, value_col: str, region_name: str = None) -> float:
    df = get_query_df(query_num)
    if df is None or len(df) == 0:
        return 0.0

    date_col = get_date_col(df)
    
    # Check if this df actually has a date column
    has_date = any(k in date_col.lower() for k in ['date', 'event'])
    
    if not has_date:
        dates = get_unique_dates()
        max_date = max(dates) if dates else date_int
        if date_int == max_date:
            try:
                return float(df[value_col].iloc[0]) if value_col in df.columns else float(df.iloc[0, 0])
            except Exception:
                try:
                    return float(df.iloc[0, 0])
                except Exception:
                    return 0.0
        return 0.0

    # Filter by date
    try:
        df_date = df[df[date_col].astype(str).str.contains(str(date_int))]
        if len(df_date) == 0:
            df_date = df[df[date_col].astype(float).astype(int) == int(date_int)]
    except Exception:
        return 0.0

    if len(df_date) == 0:
        return 0.0

    if region_name:
        region_col = None
        for col in df_date.columns:
            if 'region' in col.lower():
                region_col = col
                break
        if not region_col:
            return 0.0
        
        df_region = df_date[df_date[region_col].apply(clean_region_name) == clean_region_name(region_name)]
        if len(df_region) > 0:
            return float(df_region[value_col].iloc[0])
        return 0.0
    else:
        if len(df_date) == 1:
            try:
                return float(df_date[value_col].iloc[0])
            except (ValueError, TypeError):
                return 0.0
        else:
            try:
                return float(df_date[value_col].sum())
            except (ValueError, TypeError):
                return 0.0

def format_cell_value(val: float, is_percentage: bool = False) -> str:
    if val is None or pd.isna(val) or val == "":
        return ""
    if is_percentage:
        return f"{val * 100:.1f}%"
    try:
        return f"{int(round(val)):,}"
    except (ValueError, TypeError):
        return str(val)

def calculate_ratio(num: float, den: float) -> float:
    return num / den if den != 0 else 0.0

def generate_dashboard():
    # 1. Parse template file structure
    if not TEMPLATE_PATH.exists():
        print(f"Error: Template file not found: {TEMPLATE_PATH}")
        return

    print(f"Reading report template: {TEMPLATE_PATH}")
    with open(TEMPLATE_PATH, 'r', encoding='utf-8') as f:
        template_lines = [line.strip('\n') for line in f.readlines()]

    row_labels = []
    for line in template_lines:
        parts = line.split('\t')
        row_labels.append(parts[0])

    # 2. Load existing report to preserve history (indexed by row index to prevent region collisions)
    history_map = {} # date_label -> {row_idx -> val}
    history_headers = {} # date_label -> day_of_week
    ordered_date_labels = []
    
    if OUTPUT_REPORT_PATH.exists():
        try:
            print(f"Found existing pivoted report at {OUTPUT_REPORT_PATH}. Loading history...")
            with open(OUTPUT_REPORT_PATH, 'r', encoding='utf-8') as f:
                lines = [line.strip('\n') for line in f.readlines()]
            if len(lines) >= 3:
                days_of_week = lines[1].split('\t')[1:]
                date_labels = lines[2].split('\t')[1:]
                
                days_of_week = [d.strip() for d in days_of_week]
                date_labels = [d.strip() for d in date_labels]
                
                num_cols = min(len(days_of_week), len(date_labels))
                
                for col_idx in range(num_cols):
                    date_lbl = date_labels[col_idx]
                    ordered_date_labels.append(date_lbl)
                    history_headers[date_lbl] = days_of_week[col_idx]
                    history_map[date_lbl] = {}
                    
                for idx, line in enumerate(lines[3:]):
                    row_idx = idx + 3
                    if not line:
                        continue
                    parts = line.split('\t')
                    for col_idx in range(num_cols):
                        date_lbl = date_labels[col_idx]
                        val = parts[col_idx + 1] if len(parts) > (col_idx + 1) else ""
                        history_map[date_lbl][row_idx] = val
            print(f"Loaded {len(ordered_date_labels)} historical dates: {ordered_date_labels}")
        except Exception as e:
            print(f"Warning: Error loading history: {e}. Starting fresh.")
            history_map = {}
            history_headers = {}
            ordered_date_labels = []

    # 3. Find all unique dates in query outputs
    dates = get_unique_dates()
    if not dates:
        print("Error: No dates found in outputs/query_*.csv files. Please run the SQL queries first.")
        return
    print(f"Found dates in query outputs: {dates}")

    # 4. Generate data for query dates and update history
    regions = [
        "West Addis", "East Addis", "Central", "South", "North West", 
        "North East", "East 1", "East 2", "West", "North", "Afar", "Unknown"
    ]

    block_configs = {
        0: (3, "VLR_ATTCHED", None, None),                         # Daily Active subs
        1: (5, "SUBS", 4, "SUBS"),                                 # Recharge - Subs
        2: (5, "TRAFFIC", None, None),                             # Recharge - Amount
        3: (7, "subs", 6, "subs"),                                 # Loan Subs
        4: (7, "TRAFFIC", None, None),                             # Loan Amount
        5: (9, "SUBS", 8, "usb"),                                  # Repayment Subs
        6: (9, "TRAFFIC", None, None),                             # Repayment Amount
        7: (11, "subs", 10, "subs"),                               # Direct Bundle Purchase Subs_ MPESA
        8: (11, "Bundle_VALUE_TRAFFIC", None, None),               # Direct Bundle Purchase Amount_ MPESA
        9: (13, "SUBS", 12, "GA"),                                 # Gross adds
        10: (15, "total_Subs", 14, "total_Subs"),                  # Daily Voice active subs
        11: (15, "TRAFFIC", None, None),                           # Voice - mins
        12: (17, "total_Subs", 16, "total_Subs"),                  # Daily data active subs
        13: (17, "TRAFFIC", None, None),                           # Data - GB
        14: (19, "Subs", 18, "total_Subs"),                        # Daily SMS active subs
        15: (19, "TRAFFIC", None, None)                            # SMS count
    }

    for date_int in dates:
        day_of_week, date_label = format_date_headers(date_int)
        
        # Add to ordered list if it's a new date
        if date_label not in ordered_date_labels:
            ordered_date_labels.append(date_label)
            
        history_headers[date_label] = day_of_week
        history_map[date_label] = {}
        
        # Calculate daily active, 30 days active, 90 days active
        daily_active = extract_metric(3, date_int, "VLR_ATTCHED")
        active_30 = extract_metric(1, date_int, "count(DISTINCT msisdn)")
        active_90 = extract_metric(2, date_int, "count(DISTINCT msisdn)")
        
        daily_from_30 = calculate_ratio(daily_active, active_30) if active_30 > 0 else 0.0
        active_30_from_90 = calculate_ratio(active_30, active_90) if active_90 > 0 else 0.0
        
        # Store overview values
        history_map[date_label][3] = format_cell_value(daily_active)
        history_map[date_label][4] = format_cell_value(active_30)
        history_map[date_label][5] = format_cell_value(active_90)
        history_map[date_label][6] = format_cell_value(daily_from_30, is_percentage=True)
        history_map[date_label][7] = format_cell_value(active_30_from_90, is_percentage=True)
        
        # Store blocks
        for k, (reg_q, reg_col, nat_q, nat_col) in block_configs.items():
            base_idx = 8 + k * 13
            
            # Overall value
            if nat_q and nat_col:
                overall_val = extract_metric(nat_q, date_int, nat_col)
            else:
                overall_val = extract_metric(reg_q, date_int, reg_col)
                
            history_map[date_label][base_idx] = format_cell_value(overall_val)
            
            # Regional values
            for r_offset, region in enumerate(regions, start=1):
                idx = base_idx + r_offset
                reg_val = extract_metric(reg_q, date_int, reg_col, region_name=region)
                history_map[date_label][idx] = format_cell_value(reg_val)

    # 5. Build output columns matrix
    output_cols = []
    for date_lbl in ordered_date_labels:
        col_values = [""] * len(template_lines)
        day_of_week = history_headers.get(date_lbl, "")
        
        col_values[1] = f" {day_of_week} "
        col_values[2] = date_lbl
        
        date_data = history_map.get(date_lbl, {})
        for row_idx in range(3, len(template_lines)):
            col_values[row_idx] = date_data.get(row_idx, "")
            
        output_cols.append(col_values)

    # 6. Write output to outputs/real-sample-report-pivoted.csv
    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    
    with open(OUTPUT_REPORT_PATH, 'w', encoding='utf-8') as f:
        for row_idx in range(len(template_lines)):
            row_label = row_labels[row_idx]
            row_vals = [col[row_idx] for col in output_cols]
            f.write(f"{row_label}\t" + "\t".join(row_vals) + "\n")
            
    print(f"Daily Stand-up Dashboard pivoted report generated at: {OUTPUT_REPORT_PATH}")
    
    # 7. Write to outputs/real-sample-report-pivoted.xlsx
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        excel_path = OUTPUTS_DIR / "real-sample-report-pivoted.xlsx"
        
        # Build pandas DataFrame
        data_dict = {"Metric / Region": row_labels}
        for i, date_lbl in enumerate(ordered_date_labels):
            data_dict[date_lbl] = output_cols[i]
            
        df_out = pd.DataFrame(data_dict)
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_out.to_excel(writer, sheet_name="Daily_Dashboard", index=False)
            
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Daily_Dashboard"]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin', color="CCCCCC"),
            right=Side(style='thin', color="CCCCCC"),
            top=Side(style='thin', color="CCCCCC"),
            bottom=Side(style='thin', color="CCCCCC")
        )
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for idx, cell in enumerate(col):
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
                if idx > 0:
                    cell.border = thin_border
                    if col_letter == 'A':
                        cell.alignment = Alignment(horizontal="left")
                    else:
                        cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        wb.save(excel_path)
        print(f"Excel dashboard report formatted and saved at: {excel_path}")
        
    except ImportError:
        print("Warning: pandas or openpyxl not available. Skipping Excel dashboard output.")
    except Exception as e:
        print(f"Warning: Error creating Excel dashboard: {e}")

    print(f"The output has {len(template_lines)} rows and {len(ordered_date_labels) + 1} columns (Labels + {len(ordered_date_labels)} dates).")

if __name__ == "__main__":
    generate_dashboard()
